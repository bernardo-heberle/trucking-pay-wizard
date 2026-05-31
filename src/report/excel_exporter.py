from __future__ import annotations

import datetime
import re
from pathlib import Path

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.extract.models import Certainty, DocumentExtractionResult, ExtractedLoad
from src.report._date_parsing import parse_extracted_date
from src.report.exceptions import ReportAssemblyError

_CURRENCY_FORMAT = '$#,##0.00'
_DATE_FORMAT = 'MM/DD/YYYY'
_DAYS_FORMAT = '[=1]0" day";0" days"'

_FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

_CERTAINTY_FILLS: dict[Certainty, PatternFill] = {
    Certainty.HIGH: _FILL_GREEN,
    Certainty.REVIEW: _FILL_YELLOW,
    Certainty.NOT_FOUND: _FILL_RED,
}

# Placeholder shown for fields that do not apply to an excluded document.
_NA = "NA"

# Notes shown on the single review/exclusion rows.
_REVIEW_NO_VALUES_NOTE = (
    "Classified as financial document, but tool could not extract values - REVIEW"
)
_NONFINANCIAL_NOTE = "Not included - document contains no payment information"

# Fixed column indices (1-based).
_COL_DOCUMENT = 1
_COL_PDF_PAGE = 2
_COL_CERTAINTY = 3
_COL_DATE = 4
_COL_PAY = 5
_COL_NOTES = 6
_N_FIXED_COLS = 6

_HEADERS = ["Document", "PDF Page", "Certainty", "Date", "Pay", "Notes"]


def _fill_for_certainty(certainty: Certainty | None) -> PatternFill:
    if certainty is None:
        return _FILL_YELLOW
    return _CERTAINTY_FILLS.get(certainty, _FILL_YELLOW)


def build_excel(
    results: list[DocumentExtractionResult],
    output_path: Path,
    page_offsets: dict[str, int],
    page_limits: dict[str, int] | None = None,
    duplicate_map: dict[str, list[str]] | None = None,
) -> Path:
    """Build an Excel spreadsheet describing every document in the packet.

    Column layout: ``Document | PDF Page | Certainty | Date | Pay | Notes``.

    Rows are grouped into tiers, top to bottom:

    1. Payment documents with a parseable date — one row per extracted load,
       ordered chronologically.  Multi-load documents repeat the document name.
    2. Payment documents that are in the PDF but have no parseable date — either
       their load rows (when values were extracted) or, when nothing usable was
       extracted, a single red review row whose PDF Page shows the document's
       page range (e.g. ``33 - 38``).
    3. Excluded documents, grayed out: non-payment documents (with a note that
       they contain no payment information) followed by exact duplicates (noted
       as ``Exact duplicate of <kept filename>``).  Their PDF Page and Certainty
       show ``NA``; the Date and Pay cells are left blank so the totals formulas
       ignore them.  These documents are absent from the combined PDF.
    4. A TOTALS row at the very bottom (``=SUM`` on Pay, ``=MAX-MIN+1`` on Date),
       which reflects the payment rows only (blank cells are ignored).

    *page_offsets* maps ``source_path.name`` to the 1-indexed starting page of
    that document in the combined PDF; *page_limits* maps it to the number of
    pages it spans there.  Together they give the page range for review rows.
    Per-load PDF Page is the document's base page plus the load's own offset.

    *duplicate_map* maps a kept filename to the byte-identical filenames that
    were excluded from the pipeline; each excluded name becomes its own grayed
    row.  The kept document itself carries no duplicate note.

    Returns *output_path* on success.

    Raises:
        ReportAssemblyError: The workbook cannot be saved.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    _write_header_row(ws, _HEADERS)

    _dup_map: dict[str, list[str]] = duplicate_map or {}
    _page_limits: dict[str, int] = page_limits or {}

    payment_docs = [r for r in results if r.is_payment_document]
    nonpayment_docs = [r for r in results if not r.is_payment_document]

    # Tier 1 before tier 2: dated payment docs (chronological) then undated.
    tier1 = sorted(
        (r for r in payment_docs if _doc_earliest_date(r) is not None),
        key=lambda r: _doc_earliest_date(r),  # type: ignore[arg-type, return-value]
    )
    tier2 = [r for r in payment_docs if _doc_earliest_date(r) is None]

    data_row = 2
    parseable_date_count = 0
    for result in tier1 + tier2:
        data_row, wrote_dates = _write_payment_doc(
            ws, result, data_row, page_offsets, _page_limits
        )
        parseable_date_count += wrote_dates

    has_payment_rows = data_row > 2

    # Tier 3: excluded documents, grayed out, written above the totals row.
    for result in nonpayment_docs:
        _write_excluded_row(ws, result.source_path.name, _NONFINANCIAL_NOTE, data_row)
        data_row += 1
    for kept_name, duplicate_names in _dup_map.items():
        for duplicate_name in duplicate_names:
            _write_excluded_row(
                ws,
                duplicate_name,
                f"Exact duplicate of {kept_name}",
                data_row,
            )
            data_row += 1

    # TOTALS row sits at the very bottom.  Excluded rows leave Date/Pay blank, so
    # the SUM/MAX/MIN formulas over the full data range reflect payment rows only.
    if has_payment_rows:
        totals_row = data_row
        _write_totals_row(
            ws,
            data_start_row=2,
            data_end_row=totals_row - 1,
            totals_row=totals_row,
            has_parseable_dates=parseable_date_count > 0,
        )

    _auto_width(ws, _N_FIXED_COLS)

    try:
        wb.save(str(output_path))
    except Exception as exc:
        raise ReportAssemblyError(
            f"Failed to save Excel file '{output_path}': {exc}"
        ) from exc

    logger.info(
        "Excel report saved to '{}' — {} payment document(s), "
        "{} non-payment, {} duplicate(s)",
        output_path.name,
        len(payment_docs),
        len(nonpayment_docs),
        sum(len(v) for v in _dup_map.values()),
    )
    return output_path


def _doc_earliest_date(result: DocumentExtractionResult) -> datetime.date | None:
    """Return the earliest parseable load date in *result*, or ``None``."""
    earliest: datetime.date | None = None
    for load in result.loads:
        if load.date is not None and load.date.value:
            parsed = parse_extracted_date(load.date.value)
            if parsed is not None and (earliest is None or parsed < earliest):
                earliest = parsed
    return earliest


def _page_range_text(
    doc_name: str,
    page_offsets: dict[str, int],
    page_limits: dict[str, int],
) -> str:
    """Return the document's page range in the combined PDF (e.g. ``33 - 38``).

    Returns a single page number when the document spans one page, or ``NA``
    when the document is not present in the PDF.
    """
    start = page_offsets.get(doc_name)
    limit = page_limits.get(doc_name)
    if not isinstance(start, int) or not isinstance(limit, int) or limit < 1:
        return _NA
    end = start + limit - 1
    return str(start) if end == start else f"{start} - {end}"


def _write_payment_doc(
    ws,
    result: DocumentExtractionResult,
    data_row: int,
    page_offsets: dict[str, int],
    page_limits: dict[str, int],
) -> tuple[int, int]:
    """Write the row(s) for one payment document.

    Returns ``(next_data_row, parseable_date_count)``.  Documents with no
    extractable values (including hard extraction failures) get a single red
    review row whose PDF Page shows the document's page range.
    """
    doc_name = result.source_path.name

    if not result.has_extractable_values():
        ws.cell(row=data_row, column=_COL_DOCUMENT, value=doc_name)
        ws.cell(
            row=data_row,
            column=_COL_PDF_PAGE,
            value=_page_range_text(doc_name, page_offsets, page_limits),
        )
        cert_cell = ws.cell(
            row=data_row, column=_COL_CERTAINTY, value=Certainty.REVIEW.value
        )
        cert_cell.fill = _FILL_RED
        note = _REVIEW_NO_VALUES_NOTE
        if result.extraction_error:
            note = f"{note} ({result.extraction_error})"
        notes_cell = ws.cell(row=data_row, column=_COL_NOTES, value=note)
        notes_cell.fill = _FILL_RED
        return data_row + 1, 0

    doc_base_page = page_offsets.get(doc_name, "")
    parseable_date_count = 0
    for load in result.loads:
        ws.cell(row=data_row, column=_COL_DOCUMENT, value=doc_name)

        # Per-load PDF page: base offset + (load's source_page - 1).
        load_source_page = _load_source_page(load)
        if isinstance(doc_base_page, int) and load_source_page is not None:
            pdf_page: int | str = doc_base_page + (load_source_page - 1)
        else:
            pdf_page = doc_base_page or ""
        ws.cell(row=data_row, column=_COL_PDF_PAGE, value=pdf_page)

        load_cert = load.certainty()
        cert_cell = ws.cell(row=data_row, column=_COL_CERTAINTY, value=load_cert.value)
        cert_cell.fill = _fill_for_certainty(load_cert)

        # Date cell — write datetime on success; blank + red fill + Notes on failure.
        date_certainty: Certainty | None = None
        notes_parts: list[str] = []
        if load.date is not None:
            date_raw = load.date.value
            date_certainty = load.date.certainty
            parsed_date = parse_extracted_date(date_raw)
            if parsed_date is not None:
                date_value: datetime.datetime | None = datetime.datetime(
                    parsed_date.year, parsed_date.month, parsed_date.day
                )
                parseable_date_count += 1
            else:
                date_value = None
                notes_parts.append(f'Unparseable date: "{date_raw}"')
        else:
            date_value = None

        date_cell = ws.cell(row=data_row, column=_COL_DATE, value=date_value)
        date_cell.number_format = _DATE_FORMAT
        if date_value is not None:
            date_cell.fill = _fill_for_certainty(date_certainty)
        else:
            date_cell.fill = _FILL_RED

        # Pay cell
        pay_value: float | str | None = None
        pay_certainty: Certainty | None = None
        if load.pay is not None:
            pay_certainty = load.pay.certainty
            pay_value = _parse_pay_float(load.pay.value)
        pay_cell = ws.cell(row=data_row, column=_COL_PAY, value=pay_value)
        pay_cell.number_format = _CURRENCY_FORMAT
        pay_cell.fill = (
            _fill_for_certainty(pay_certainty) if pay_value is not None else _FILL_RED
        )

        # Notes cell — join any note fragments.
        if notes_parts:
            notes_cell = ws.cell(
                row=data_row, column=_COL_NOTES, value="; ".join(notes_parts)
            )
            notes_cell.fill = _FILL_RED

        data_row += 1

    return data_row, parseable_date_count


def _write_excluded_row(ws, doc_name: str, note: str, row: int) -> None:
    """Write a grayed-out row for an excluded document.

    PDF Page and Certainty show ``NA``; the Date and Pay cells are left blank so
    the totals ``SUM``/``MAX``/``MIN`` formulas over the full data range ignore
    them.  The whole row is filled gray.
    """
    ws.cell(row=row, column=_COL_DOCUMENT, value=doc_name)
    ws.cell(row=row, column=_COL_PDF_PAGE, value=_NA)
    ws.cell(row=row, column=_COL_CERTAINTY, value=_NA)
    # Date and Pay intentionally left blank (not "NA") so totals formulas ignore them.
    ws.cell(row=row, column=_COL_DATE)
    ws.cell(row=row, column=_COL_PAY)
    ws.cell(row=row, column=_COL_NOTES, value=note)
    for col in range(1, _N_FIXED_COLS + 1):
        ws.cell(row=row, column=col).fill = _FILL_GRAY


def _load_source_page(load: ExtractedLoad) -> int | None:
    """Return the source page for *load*, preferring pay's page over date's."""
    if load.pay is not None and load.pay.source_page is not None:
        return load.pay.source_page
    if load.date is not None and load.date.source_page is not None:
        return load.date.source_page
    return None


def _write_totals_row(
    ws,
    data_start_row: int,
    data_end_row: int,
    totals_row: int,
    has_parseable_dates: bool = True,
) -> None:
    """Write a summary Totals row using Excel formulas.

    The Pay column gets a SUM formula.  The Date column gets a MAX-MIN+1
    formula when at least one data row has a parseable date; otherwise the
    Date cell is left blank to avoid a misleading ``1 day`` from an all-empty
    date range.  All cells in the totals row are bold.
    """
    bold = Font(bold=True)

    label_cell = ws.cell(row=totals_row, column=_COL_DOCUMENT, value="TOTALS")
    label_cell.font = bold

    pay_col_letter = get_column_letter(_COL_PAY)
    pay_range = f"{pay_col_letter}{data_start_row}:{pay_col_letter}{data_end_row}"
    pay_cell = ws.cell(row=totals_row, column=_COL_PAY, value=f"=SUM({pay_range})")
    pay_cell.font = bold
    pay_cell.number_format = _CURRENCY_FORMAT

    date_col_letter = get_column_letter(_COL_DATE)
    if has_parseable_dates:
        date_range = f"{date_col_letter}{data_start_row}:{date_col_letter}{data_end_row}"
        date_cell = ws.cell(
            row=totals_row, column=_COL_DATE,
            value=f"=MAX({date_range})-MIN({date_range})+1",
        )
        date_cell.font = bold
        date_cell.number_format = _DAYS_FORMAT
    else:
        date_cell = ws.cell(row=totals_row, column=_COL_DATE, value=None)
        date_cell.font = bold


def _parse_pay_float(value: str) -> float | str:
    """Convert a pay string to float for proper Excel numeric handling.

    Strips currency symbols, commas, and surrounding whitespace so that raw
    LLM values like ``'$1,500.00'`` parse correctly.  Returns the float on
    success so Excel treats the cell as a number and SUM() formulas work.
    Returns the original string unchanged when parsing fails so the cell
    still shows something reviewable.
    """
    try:
        return float(re.sub(r"[$,\s]", "", value))
    except (ValueError, AttributeError):
        return value


def _write_header_row(ws, headers: list[str]) -> None:
    bold = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws, num_columns: int) -> None:
    """Set column widths based on the longest value in each column."""
    for col in range(1, num_columns + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)
