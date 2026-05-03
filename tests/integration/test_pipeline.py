"""Integration test: Ingest -> mocked OCR -> mocked LLM Extract -> Report.

Exercises the full pipeline with synthetic PDF documents and a mocked Azure
client and mocked Anthropic client, verifying that extracted values flow
correctly through all stages and appear in the final Excel and combined PDF
outputs.
"""

from __future__ import annotations

import datetime
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

import fitz
import openpyxl
import pytest

from src.config import Settings
from src.extract.llm.extractor import LlmExtractor
from src.extract.models import Certainty, DocumentExtractionResult, ExtractedField, ExtractedLoad, SourceSpan
from src.ingest import ingest_document
from src.ocr.models import BoundingBox, OcrLine, OcrPage, OcrResult
from src.report import build_report


def _make_synthetic_pdf(path: Path, text_lines: list[tuple[str, float]]) -> None:
    """Create a single-page PDF with *text_lines* at specified y positions."""
    doc = fitz.open()
    page = doc.new_page(width=612, height=792)
    for text, y in text_lines:
        page.insert_text((72, y), text)
    doc.save(str(path))
    doc.close()


def _build_mock_ocr_result(source_path: Path, content_hash: str) -> OcrResult:
    """Build a canned OcrResult for use with the mocked extractor."""
    lines_text = [
        "Settlement Statement",
        "Order ID: BSAT1066",
        "Total Payment to Carrier: $1,200.50",
        "Pickup Exactly: 05/15/2024",
    ]
    lines: list[OcrLine] = []
    offset = 0
    for i, text in enumerate(lines_text):
        start = offset
        end = offset + len(text)
        lines.append(
            OcrLine(
                text=text,
                page_number=1,
                bounding_box=BoundingBox(x=1.0, y=0.5 + i * 1.5, width=4.0, height=0.25),
                char_start=start,
                char_end=end,
            )
        )
        offset = end + (1 if i < len(lines_text) - 1 else 0)

    return OcrResult(
        source_path=source_path,
        content_hash=content_hash,
        pages=[OcrPage(page_number=1, width_inches=8.5, height_inches=11.0, line_count=len(lines))],
        lines=lines,
    )


def _make_extraction_result(source_path: Path, content_hash: str, page_count: int):
    """Build a canned DocumentExtractionResult as if returned by LlmExtractor."""
    pay = ExtractedField(
        name="pay",
        value="1,200.50",
        source_document=source_path.name,
        source_page=1,
        source_spans=[SourceSpan(
            page_number=1,
            bounding_box=BoundingBox(x=1.0, y=2.0, width=4.0, height=0.25),
        )],
        certainty=Certainty.HIGH,
        confidence=0.97,
    )
    date = ExtractedField(
        name="date",
        value="05/15/2024",
        source_document=source_path.name,
        source_page=1,
        source_spans=[SourceSpan(
            page_number=1,
            bounding_box=BoundingBox(x=1.0, y=3.5, width=3.5, height=0.25),
        )],
        certainty=Certainty.HIGH,
        confidence=0.95,
    )
    return DocumentExtractionResult(
        source_path=source_path,
        content_hash=content_hash,
        page_count=page_count,
        loads=[ExtractedLoad(index=1, pay=pay, date=date)],
    )


class TestFullPipeline:

    def test_end_to_end(self, tmp_path: Path) -> None:
        # 1. Create a synthetic source PDF
        source = tmp_path / "test_settlement.pdf"
        _make_synthetic_pdf(source, [
            ("Settlement Statement", 72),
            ("Total Payment to Carrier: $1,200.50", 200),
            ("Pickup Exactly: 05/15/2024", 320),
        ])

        # 2. Ingest
        ingested = ingest_document(source)
        assert ingested.page_count >= 1

        # 3. Build a mock OCR result (bypassing Azure)
        ocr_result = _build_mock_ocr_result(source, ingested.content_hash)

        # 4. Extract using a mocked LlmExtractor
        canned = _make_extraction_result(source, ingested.content_hash, ingested.page_count)
        mock_extractor = MagicMock(spec=LlmExtractor)
        mock_extractor.extract.return_value = canned

        extraction = mock_extractor.extract(ocr_result, page_count=ingested.page_count)
        assert len(extraction.loads) == 1

        pay = extraction.loads[0].pay
        assert pay is not None
        assert pay.value == "1,200.50"
        assert pay.certainty == Certainty.HIGH

        date = extraction.loads[0].date
        assert date is not None
        assert date.value == "05/15/2024"
        assert date.certainty == Certainty.HIGH

        # 5. Report assembly
        output_dir = tmp_path / "output"
        pdf_path, excel_path = build_report([extraction], output_dir)

        assert pdf_path.exists()
        assert excel_path.exists()

        # Verify PDF page count: source pages only (no index page)
        combined = fitz.open(str(pdf_path))
        assert len(combined) == ingested.page_count

        # Verify highlight annotations on source page(s)
        source_page = combined[0]
        annots = list(source_page.annots() or [])
        assert len(annots) >= 1
        combined.close()

        # Verify Excel values
        wb = openpyxl.load_workbook(str(excel_path))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        # Pay is stored as float so Excel SUM formulas work.
        assert ws.cell(row=2, column=header_map["Pay"]).value == pytest.approx(1200.50)
        # Date is stored as a datetime object so Excel MAX/MIN formulas work.
        assert ws.cell(row=2, column=header_map["Date"]).value == datetime.datetime(2024, 5, 15)
        assert ws.cell(row=2, column=header_map["PDF Page"]).value == 1

        assert "Certainty" in header_map
        assert ws.cell(row=2, column=header_map["Certainty"]).value == "High"

    def test_multiple_documents(self, tmp_path: Path) -> None:
        """Verify the pipeline handles multiple documents correctly."""
        extractions = []

        for idx, (pay_val, date_val) in enumerate([("500.00", "01/01/2024"), ("900.00", "06/30/2024")]):
            src = tmp_path / f"doc_{idx}.pdf"
            _make_synthetic_pdf(src, [
                ("Settlement", 72),
                (f"Total Payment to Carrier: ${pay_val}", 200),
                (f"Pickup Exactly: {date_val}", 320),
            ])

            ingested = ingest_document(src)
            pay_field = ExtractedField(
                name="pay",
                value=pay_val,
                source_document=src.name,
                source_page=1,
                source_spans=[SourceSpan(
                    page_number=1,
                    bounding_box=BoundingBox(x=1.0, y=2.0, width=4.0, height=0.25),
                )],
                certainty=Certainty.HIGH,
                confidence=0.95,
            )
            date_field = ExtractedField(
                name="date",
                value=date_val,
                source_document=src.name,
                source_page=1,
                source_spans=[SourceSpan(
                    page_number=1,
                    bounding_box=BoundingBox(x=1.0, y=3.5, width=3.5, height=0.25),
                )],
                certainty=Certainty.HIGH,
                confidence=0.95,
            )
            canned = DocumentExtractionResult(
                source_path=src,
                content_hash=ingested.content_hash,
                page_count=1,
                loads=[ExtractedLoad(index=1, pay=pay_field, date=date_field)],
            )
            extractions.append(canned)

        output_dir = tmp_path / "output"
        pdf_path, excel_path = build_report(extractions, output_dir)

        combined = fitz.open(str(pdf_path))
        assert len(combined) == 2  # 2 source pages, no index
        combined.close()

        wb = openpyxl.load_workbook(str(excel_path))
        ws = wb.active
        assert ws.max_row == 4  # header + 2 data rows + 1 totals row


def _make_settings_for_integration() -> Settings:
    return Settings(
        anthropic_api_key="sk-test-key",
        llm_model="claude-3-5-haiku-20241022",
        confidence_high_threshold=0.9,
        confidence_review_threshold=0.6,
    )


def _mock_tool_response(tool_name: str, tool_input: dict) -> MagicMock:
    tool_block = SimpleNamespace(type="tool_use", name=tool_name, input=tool_input)
    response = MagicMock()
    response.content = [tool_block]
    return response


class TestRealLlmExtractorWithMockedClient:
    """Integration tests using a real LlmExtractor (not a Mock) with a mocked
    Anthropic client.

    The existing TestFullPipeline mocks the extractor wholesale, meaning
    schema parsing, PII sanitization, pay verification, and certainty
    downgrade are never exercised.  These tests fix that gap by injecting
    a mock only at the network boundary (client.messages.create).
    """

    def _make_ocr_result(self, source_path: Path, content_hash: str) -> OcrResult:
        lines_text = [
            "Settlement Statement",
            "Total Payment to Carrier: $750.00",
            "SSN: 123-45-6789",  # PII — must be stripped before API call
            "Pickup Date: 03/12/2024",
        ]
        lines: list[OcrLine] = []
        offset = 0
        for i, text in enumerate(lines_text):
            lines.append(
                OcrLine(
                    text=text,
                    page_number=1,
                    bounding_box=BoundingBox(x=1.0, y=0.5 + i * 1.5, width=4.0, height=0.25),
                    char_start=offset,
                    char_end=offset + len(text),
                )
            )
            offset += len(text) + 1

        return OcrResult(
            source_path=source_path,
            content_hash=content_hash,
            pages=[OcrPage(page_number=1, width_inches=8.5, height_inches=11.0, line_count=len(lines))],
            lines=lines,
        )

    @patch("src.extract.llm.extractor.time.sleep")
    def test_real_extractor_parses_schema_and_verifies_pay(
        self, _mock_sleep, tmp_path: Path
    ) -> None:
        """The real LlmExtractor must parse the schema, sanitize PII, and
        verify the pay value against OCR text — producing HIGH certainty when
        the value is found.
        """
        source = tmp_path / "settlement.pdf"
        doc = fitz.open()
        doc.new_page(width=612, height=792)
        doc.save(str(source))
        doc.close()

        ingested = ingest_document(source)
        ocr = self._make_ocr_result(source, ingested.content_hash)

        mock_client = MagicMock()
        mock_client.messages.create.return_value = _mock_tool_response(
            "extract_income_fields",
            {
                "pay": {"value": "750.00", "confidence": 0.95},
                "date": {"value": "03/12/2024", "confidence": 0.92},
            },
        )

        settings = _make_settings_for_integration()
        extractor = LlmExtractor(client=mock_client, settings=settings)
        result = extractor.extract(ocr, page_count=ingested.page_count)

        # Schema parsing produced correctly structured loads.
        assert result.extraction_error is None
        assert len(result.loads) == 1

        pay = result.loads[0].pay
        assert pay is not None
        assert pay.value == "750.00"
        # OCR text contains "$750.00" — verification must pass and keep HIGH.
        assert pay.certainty == Certainty.HIGH

        date = result.loads[0].date
        assert date is not None
        assert date.value == "03/12/2024"
        assert date.certainty == Certainty.HIGH

        # PII (SSN) must have been stripped before the API call.
        call_args = mock_client.messages.create.call_args
        user_message = call_args.kwargs["messages"][0]["content"]
        assert "123-45-6789" not in user_message
        assert "[SSN-REDACTED]" in user_message

    @patch("src.extract.llm.extractor.time.sleep")
    def test_real_extractor_downgrades_certainty_when_pay_not_in_ocr(
        self, _mock_sleep, tmp_path: Path
    ) -> None:
        """When the LLM returns a pay value that does not appear in the OCR
        text, the real _verify_pay_fields must downgrade certainty to REVIEW.
        Mocking the extractor wholesale would hide this behavior.
        """
        source = tmp_path / "settlement.pdf"
        doc = fitz.open()
        doc.new_page(width=612, height=792)
        doc.save(str(source))
        doc.close()

        ingested = ingest_document(source)
        # OCR text contains $750.00, not $9999.00
        ocr = self._make_ocr_result(source, ingested.content_hash)

        mock_client = MagicMock()
        mock_client.messages.create.return_value = _mock_tool_response(
            "extract_income_fields",
            {
                # LLM returns a wrong value with high confidence.
                "pay": {"value": "9999.00", "confidence": 0.95},
                "date": None,
            },
        )

        settings = _make_settings_for_integration()
        extractor = LlmExtractor(client=mock_client, settings=settings)
        result = extractor.extract(ocr, page_count=ingested.page_count)

        assert result.extraction_error is None
        assert len(result.loads) == 1
        pay = result.loads[0].pay
        assert pay is not None
        assert pay.value == "9999.00"
        # Value not found in OCR → must be downgraded from HIGH to REVIEW.
        assert pay.certainty == Certainty.REVIEW

    @patch("src.extract.llm.extractor.time.sleep")
    def test_real_extractor_preserves_raw_pay_and_resolves_location(
        self, _mock_sleep, tmp_path: Path
    ) -> None:
        """The LLM now returns the value exactly as it appears in the document.
        The schema stores it verbatim; the resolver finds it in the OCR text;
        and pay verification normalizes it before comparing against OCR amounts.
        """
        source = tmp_path / "settlement.pdf"
        doc = fitz.open()
        doc.new_page(width=612, height=792)
        doc.save(str(source))
        doc.close()

        ingested = ingest_document(source)
        ocr_text = "Total Payment: $1,500.00"
        lines = [
            OcrLine(
                text=ocr_text,
                page_number=1,
                bounding_box=BoundingBox(x=1.0, y=1.0, width=4.0, height=0.25),
                char_start=0,
                char_end=len(ocr_text),
            )
        ]
        ocr = OcrResult(
            source_path=source,
            content_hash=ingested.content_hash,
            pages=[OcrPage(page_number=1, width_inches=8.5, height_inches=11.0, line_count=1)],
            lines=lines,
        )

        mock_client = MagicMock()
        # LLM returns the formatted value as it appears in the document.
        mock_client.messages.create.return_value = _mock_tool_response(
            "extract_income_fields",
            {
                "pay": {"value": "$1,500.00", "confidence": 0.95},
                "date": None,
            },
        )

        settings = _make_settings_for_integration()
        extractor = LlmExtractor(client=mock_client, settings=settings)
        result = extractor.extract(ocr, page_count=1)

        assert len(result.loads) == 1
        pay = result.loads[0].pay
        assert pay is not None
        # Raw value is preserved verbatim — normalization is deferred to Excel export.
        assert pay.value == "$1,500.00"
        # OCR text contains "$1,500.00" — verification must pass and keep HIGH.
        assert pay.certainty == Certainty.HIGH
        # Source location must be resolved from the OCR line.
        assert len(pay.source_spans) == 1
        assert pay.source_page == 1
