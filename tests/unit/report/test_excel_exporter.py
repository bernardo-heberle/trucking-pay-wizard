"""Unit tests for src.report.excel_exporter."""

from __future__ import annotations

import datetime
from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils import get_column_letter

from src.extract.models import (
    Certainty,
    DocumentExtractionResult,
    ExtractedField,
    ExtractedLoad,
)
from src.report.excel_exporter import (
    build_excel,
    _DAYS_FORMAT,
    _FILL_GREEN,
    _FILL_YELLOW,
    _FILL_RED,
    _NONFINANCIAL_NOTE,
    _REVIEW_NO_VALUES_NOTE,
)
from tests.unit.report.conftest import make_extraction_result


class TestExcelStructure:

    def test_header_row(self, synthetic_source_pdf: Path, tmp_path: Path) -> None:
        result = make_extraction_result(synthetic_source_pdf)
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert headers[0] == "Document"
        assert headers[1] == "PDF Page"
        assert headers[2] == "Certainty"
        assert "Pay" in headers
        assert "Date" in headers

    def test_row_count_matches_results(
        self, synthetic_source_pdf: Path, synthetic_source_pdf_b: Path, tmp_path: Path
    ) -> None:
        results = [
            make_extraction_result(synthetic_source_pdf, content_hash="h1"),
            make_extraction_result(synthetic_source_pdf_b, content_hash="h2", page_count=2),
        ]
        offsets = {synthetic_source_pdf.name: 2, synthetic_source_pdf_b.name: 3}
        out = tmp_path / "out.xlsx"
        build_excel(results, out, offsets)

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        assert ws.max_row == 4  # 1 header + 2 data + 1 totals

    def test_field_values_in_correct_columns(self, synthetic_source_pdf: Path, tmp_path: Path) -> None:
        result = make_extraction_result(synthetic_source_pdf)
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active

        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        gp_col = header_map["Pay"]
        dd_col = header_map["Date"]

        assert ws.cell(row=2, column=gp_col).value == pytest.approx(750.0)
        # Dates are stored as datetime objects so Excel MAX/MIN formulas work.
        assert ws.cell(row=2, column=dd_col).value == datetime.datetime(2024, 3, 12)

    def test_pdf_page_offset(self, synthetic_source_pdf: Path, tmp_path: Path) -> None:
        result = make_extraction_result(synthetic_source_pdf)
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 5})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        assert ws.cell(row=2, column=2).value == 5


class TestFixedColumnLayout:
    """The new exporter uses a fixed column set: Document | PDF Page | Certainty | Date | Pay | Notes."""

    def test_notes_column_present(self, synthetic_source_pdf: Path, tmp_path: Path) -> None:
        result = make_extraction_result(synthetic_source_pdf)
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Notes" in headers

    def test_no_date_field_leaves_date_cell_empty(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        """A load with date=None must produce an empty Date cell (not raise)."""
        result = make_extraction_result(
            synthetic_source_pdf,
            fields=[
                ExtractedField(name="pay", value="100", source_document="a.pdf", source_page=1),
            ],
        )
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        assert ws.cell(row=2, column=header_map["Date"]).value is None

    def test_multiple_loads_produce_multiple_rows(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        """A three-load document must produce three data rows."""
        from src.extract.models import ExtractedLoad
        from tests.unit.report.conftest import make_load

        loads = [
            make_load(synthetic_source_pdf, index=i, pay_value=f"{i * 100}.00", date_value=f"0{i}/01/2024")
            for i in range(1, 4)
        ]
        result = make_extraction_result(synthetic_source_pdf, loads=loads)
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        # 1 header + 3 data rows + 1 totals = 5
        assert ws.max_row == 5

    def test_document_name_repeated_for_all_loads(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        """All load rows for the same document must repeat the document name."""
        from tests.unit.report.conftest import make_load

        loads = [
            make_load(synthetic_source_pdf, index=i, pay_value=f"{500 + i * 100}.00")
            for i in range(1, 3)
        ]
        result = make_extraction_result(synthetic_source_pdf, loads=loads)
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == synthetic_source_pdf.name
        assert ws.cell(row=3, column=1).value == synthetic_source_pdf.name


class TestDateNormalization:
    """Date cells are written as datetime objects (not strings) so Excel MAX/MIN work."""

    def test_month_name_date_written_as_datetime(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        """'March 20, 2024' is stored as datetime(2024, 3, 20) for native date handling."""
        result = make_extraction_result(
            synthetic_source_pdf,
            fields=[
                ExtractedField(
                    name="date", value="March 20, 2024",
                    source_document=synthetic_source_pdf.name, source_page=1,
                ),
            ],
        )
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        date_col = header_map["Date"]
        assert ws.cell(row=2, column=date_col).value == datetime.datetime(2024, 3, 20)

    def test_abbreviated_month_date_written_as_datetime(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        """'Mar 11, 2024' (abbreviated) is stored as datetime(2024, 3, 11)."""
        result = make_extraction_result(
            synthetic_source_pdf,
            fields=[
                ExtractedField(
                    name="date", value="Mar 11, 2024",
                    source_document=synthetic_source_pdf.name, source_page=1,
                ),
            ],
        )
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        date_col = header_map["Date"]
        assert ws.cell(row=2, column=date_col).value == datetime.datetime(2024, 3, 11)

    def test_numeric_date_written_as_datetime(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        """A date already in MM/DD/YYYY format is stored as datetime(2024, 5, 15)."""
        result = make_extraction_result(
            synthetic_source_pdf,
            fields=[
                ExtractedField(
                    name="date", value="05/15/2024",
                    source_document=synthetic_source_pdf.name, source_page=1,
                ),
            ],
        )
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        date_col = header_map["Date"]
        assert ws.cell(row=2, column=date_col).value == datetime.datetime(2024, 5, 15)

    def test_unparseable_date_blanks_cell_and_writes_to_notes(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        """Unparseable date: Date cell is None+red; raw value appears in Notes."""
        result = make_extraction_result(
            synthetic_source_pdf,
            fields=[
                ExtractedField(
                    name="date", value="unknown date",
                    source_document=synthetic_source_pdf.name, source_page=1,
                    certainty=Certainty.HIGH,
                ),
                ExtractedField(
                    name="pay", value="500.00",
                    source_document=synthetic_source_pdf.name, source_page=1,
                    certainty=Certainty.HIGH,
                ),
            ],
        )
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

        date_cell = ws.cell(row=2, column=header_map["Date"])
        assert date_cell.value is None, (
            f"Expected None in Date cell for unparseable input, got {date_cell.value!r}"
        )
        assert date_cell.fill.start_color.rgb == "00FFC7CE", (
            "Expected red fill on unparseable Date cell"
        )

        notes_val = ws.cell(row=2, column=header_map["Notes"]).value
        assert notes_val is not None, "Expected Notes cell to contain the raw date string"
        assert 'Unparseable date: "unknown date"' in notes_val


class TestCertaintyColumn:

    def test_certainty_column_present(self, synthetic_source_pdf: Path, tmp_path: Path) -> None:
        result = make_extraction_result(synthetic_source_pdf)
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        assert "Certainty" in header_map

    def test_certainty_value_all_high(self, synthetic_source_pdf: Path, tmp_path: Path) -> None:
        """Default fixture has both fields HIGH -> overall is 'High'."""
        result = make_extraction_result(synthetic_source_pdf)
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        cert_col = header_map["Certainty"]
        assert ws.cell(row=2, column=cert_col).value == "High"

    def test_certainty_value_mixed(self, synthetic_source_pdf: Path, tmp_path: Path) -> None:
        result = make_extraction_result(
            synthetic_source_pdf,
            fields=[
                ExtractedField(
                    name="pay", value="100", source_document="t.pdf",
                    source_page=1, certainty=Certainty.HIGH,
                ),
                ExtractedField(
                    name="date", value="01/01/2024", source_document="t.pdf",
                    source_page=1, certainty=Certainty.REVIEW,
                ),
            ],
        )
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        cert_col = header_map["Certainty"]
        assert ws.cell(row=2, column=cert_col).value == "Review"

    def test_certainty_fill_colors(self, synthetic_source_pdf: Path, tmp_path: Path) -> None:
        """Data cells and certainty cell should carry the right fill color."""
        result = make_extraction_result(synthetic_source_pdf)
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

        cert_cell = ws.cell(row=2, column=header_map["Certainty"])
        assert cert_cell.fill.start_color.rgb == "00C6EFCE"

        pay_cell = ws.cell(row=2, column=header_map["Pay"])
        assert pay_cell.fill.start_color.rgb == "00C6EFCE"

class TestTotalsRow:

    def _build(self, tmp_path: Path, results, offsets=None) -> "openpyxl.Workbook":
        out = tmp_path / "out.xlsx"
        if offsets is None:
            offsets = {r.source_path.name: i + 2 for i, r in enumerate(results)}
        build_excel(results, out, offsets)
        # data_only=False so formula strings are preserved (not evaluated)
        return openpyxl.load_workbook(str(out), data_only=False)

    def test_totals_row_exists(
        self, synthetic_source_pdf: Path, synthetic_source_pdf_b: Path, tmp_path: Path
    ) -> None:
        """1 header + 2 data rows + 1 totals = 4 rows total."""
        results = [
            make_extraction_result(synthetic_source_pdf, content_hash="h1"),
            make_extraction_result(synthetic_source_pdf_b, content_hash="h2", page_count=2),
        ]
        wb = self._build(tmp_path, results)
        ws = wb.active
        assert ws.max_row == 4

    def test_totals_label(self, synthetic_source_pdf: Path, tmp_path: Path) -> None:
        result = make_extraction_result(synthetic_source_pdf)
        wb = self._build(tmp_path, [result])
        ws = wb.active
        totals_row = ws.max_row
        assert ws.cell(row=totals_row, column=1).value == "TOTALS"

    def test_totals_row_is_bold(self, synthetic_source_pdf: Path, tmp_path: Path) -> None:
        result = make_extraction_result(synthetic_source_pdf)
        wb = self._build(tmp_path, [result])
        ws = wb.active
        totals_row = ws.max_row
        assert ws.cell(row=totals_row, column=1).font.bold is True

    def test_pay_column_has_sum_formula(self, synthetic_source_pdf: Path, tmp_path: Path) -> None:
        result = make_extraction_result(synthetic_source_pdf)
        wb = self._build(tmp_path, [result])
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        totals_row = ws.max_row
        pay_col_letter = get_column_letter(header_map["Pay"])
        pay_cell_value = ws.cell(row=totals_row, column=header_map["Pay"]).value
        assert isinstance(pay_cell_value, str)
        assert pay_cell_value.upper().startswith("=SUM(")
        # Pin the exact range — must span exactly the data rows.
        expected_range = f"{pay_col_letter}2:{pay_col_letter}{totals_row - 1}"
        assert expected_range in pay_cell_value

    def test_date_column_has_total_days_formula(self, synthetic_source_pdf: Path, tmp_path: Path) -> None:
        """Date totals must use MAX-MIN+1 to compute inclusive calendar day span."""
        result = make_extraction_result(synthetic_source_pdf)
        wb = self._build(tmp_path, [result])
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        totals_row = ws.max_row
        date_col = header_map["Date"]
        date_col_letter = get_column_letter(date_col)
        date_cell_value = ws.cell(row=totals_row, column=date_col).value
        assert isinstance(date_cell_value, str)
        assert date_cell_value.upper().startswith("=MAX(")
        assert "-MIN(" in date_cell_value.upper()
        assert "+1" in date_cell_value
        # Pin the exact data range — must span exactly the data rows.
        expected_range = f"{date_col_letter}2:{date_col_letter}{totals_row - 1}"
        assert expected_range in date_cell_value

    def test_date_totals_cell_has_days_format(self, synthetic_source_pdf: Path, tmp_path: Path) -> None:
        """The days total must use the singular/plural-aware custom number format."""
        result = make_extraction_result(synthetic_source_pdf)
        wb = self._build(tmp_path, [result])
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        totals_row = ws.max_row
        fmt = ws.cell(row=totals_row, column=header_map["Date"]).number_format
        assert fmt == _DAYS_FORMAT

    def test_non_formula_columns_blank_in_totals(self, synthetic_source_pdf: Path, tmp_path: Path) -> None:
        """Certainty, PDF Page, and Notes columns should be empty in the totals row."""
        result = make_extraction_result(synthetic_source_pdf)
        wb = self._build(tmp_path, [result])
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        totals_row = ws.max_row
        for col_name in ("PDF Page", "Certainty", "Notes"):
            val = ws.cell(row=totals_row, column=header_map[col_name]).value
            assert val is None or val == "", f"Expected blank in '{col_name}' totals cell, got {val!r}"

    def test_sum_formula_covers_correct_range_three_docs(
        self,
        synthetic_source_pdf: Path,
        synthetic_source_pdf_b: Path,
        tmp_path: Path,
    ) -> None:
        """SUM range must span exactly the data rows (rows 2 through N)."""
        import fitz as _fitz

        pdf_c = tmp_path / "source_c.pdf"
        doc = _fitz.open()
        doc.new_page(width=612, height=792)
        doc.save(str(pdf_c))
        doc.close()

        results = [
            make_extraction_result(synthetic_source_pdf, content_hash="h1"),
            make_extraction_result(synthetic_source_pdf_b, content_hash="h2", page_count=2),
            make_extraction_result(pdf_c, content_hash="h3"),
        ]
        wb = self._build(tmp_path, results)
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        totals_row = ws.max_row  # should be 5 (1 header + 3 data + 1 totals)
        pay_col_letter = get_column_letter(header_map["Pay"])
        pay_formula = ws.cell(row=totals_row, column=header_map["Pay"]).value
        expected_range = f"{pay_col_letter}2:{pay_col_letter}{totals_row - 1}"
        assert expected_range in pay_formula

    def test_single_document_gets_totals(self, synthetic_source_pdf: Path, tmp_path: Path) -> None:
        """Edge case: 1 document still gets a totals row with valid formulas."""
        result = make_extraction_result(synthetic_source_pdf)
        wb = self._build(tmp_path, [result])
        ws = wb.active
        # 1 header + 1 data + 1 totals = 3 rows
        assert ws.max_row == 3
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        pay_col_letter = get_column_letter(header_map["Pay"])
        pay_formula = ws.cell(row=3, column=header_map["Pay"]).value
        assert f"{pay_col_letter}2:{pay_col_letter}2" in pay_formula

    def test_pay_totals_cell_has_currency_format(self, synthetic_source_pdf: Path, tmp_path: Path) -> None:
        result = make_extraction_result(synthetic_source_pdf)
        wb = self._build(tmp_path, [result])
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        totals_row = ws.max_row
        fmt = ws.cell(row=totals_row, column=header_map["Pay"]).number_format
        assert "$" in fmt

    def test_total_days_cell_uses_singular_plural_format(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        """Totals Date cell must use the conditional day/days format — not plain '0'."""
        result = make_extraction_result(synthetic_source_pdf)
        wb = self._build(tmp_path, [result])
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        totals_row = ws.max_row
        fmt = ws.cell(row=totals_row, column=header_map["Date"]).number_format
        assert fmt == _DAYS_FORMAT

    def test_total_days_cell_value_is_still_a_formula(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        """The days total must be a live formula, not a string with '& \" days\"'."""
        result = make_extraction_result(synthetic_source_pdf)
        wb = self._build(tmp_path, [result])
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        totals_row = ws.max_row
        cell_value = ws.cell(row=totals_row, column=header_map["Date"]).value
        assert isinstance(cell_value, str), f"Expected formula string, got {type(cell_value)}"
        assert cell_value.upper().startswith("=MAX("), (
            f"Expected MAX formula, got: {cell_value!r}"
        )

    def test_total_days_cell_blank_when_no_parseable_dates(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        """When every load has an unparseable date the totals Date cell must be blank."""
        result = make_extraction_result(
            synthetic_source_pdf,
            fields=[
                ExtractedField(
                    name="pay", value="500.00",
                    source_document=synthetic_source_pdf.name, source_page=1,
                    certainty=Certainty.HIGH,
                ),
                ExtractedField(
                    name="date", value="garbage",
                    source_document=synthetic_source_pdf.name, source_page=1,
                    certainty=Certainty.HIGH,
                ),
            ],
        )
        wb = self._build(tmp_path, [result])
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        totals_row = ws.max_row
        date_totals_val = ws.cell(row=totals_row, column=header_map["Date"]).value
        assert date_totals_val is None or date_totals_val == "", (
            f"Expected blank totals Date cell when no parseable dates, got {date_totals_val!r}"
        )
        # Pay totals formula must still be present.
        pay_totals_val = ws.cell(row=totals_row, column=header_map["Pay"]).value
        assert isinstance(pay_totals_val, str) and pay_totals_val.upper().startswith("=SUM(")

    def test_missing_field_gets_red_fill(
        self, synthetic_source_pdf: Path, synthetic_source_pdf_b: Path, tmp_path: Path
    ) -> None:
        """When a field is absent for a document, its cell should be red."""
        r1 = make_extraction_result(
            synthetic_source_pdf,
            content_hash="h1",
            fields=[
                ExtractedField(name="pay", value="100", source_document="a.pdf", source_page=1, certainty=Certainty.HIGH),
            ],
        )
        r2 = make_extraction_result(
            synthetic_source_pdf_b,
            content_hash="h2",
            page_count=2,
            fields=[
                ExtractedField(name="pay", value="200", source_document="b.pdf", source_page=1, certainty=Certainty.HIGH),
                ExtractedField(name="date", value="01/01/2024", source_document="b.pdf", source_page=1, certainty=Certainty.HIGH),
            ],
        )
        offsets = {synthetic_source_pdf.name: 2, synthetic_source_pdf_b.name: 3}
        out = tmp_path / "out.xlsx"
        build_excel([r1, r2], out, offsets)

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

        # r1 (pay only, no date) may be ordered after r2 because it lacks a date,
        # so locate it by document name rather than assuming a fixed row.
        doc_col = header_map["Document"]
        r1_row = next(
            row
            for row in range(2, ws.max_row + 1)
            if ws.cell(row=row, column=doc_col).value == synthetic_source_pdf.name
        )
        date_cell_r1 = ws.cell(row=r1_row, column=header_map["Date"])
        assert date_cell_r1.fill.start_color.rgb == "00FFC7CE"


class TestPayNumericCells:
    """Pay cell values must be numeric (float) so Excel SUM formulas work."""

    def test_pay_cell_is_float_not_string(self, synthetic_source_pdf: Path, tmp_path: Path) -> None:
        """A normalised pay string '750.00' must be written as float 750.0."""
        result = make_extraction_result(
            synthetic_source_pdf,
            fields=[
                ExtractedField(
                    name="pay", value="750.00",
                    source_document=synthetic_source_pdf.name, source_page=1,
                    certainty=Certainty.HIGH,
                ),
            ],
        )
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        pay_cell = ws.cell(row=2, column=header_map["Pay"])

        assert isinstance(pay_cell.value, (int, float)), (
            f"Expected numeric type, got {type(pay_cell.value).__name__!r}: {pay_cell.value!r}"
        )

    def test_pay_cell_value_matches_extracted_amount(self, synthetic_source_pdf: Path, tmp_path: Path) -> None:
        result = make_extraction_result(
            synthetic_source_pdf,
            fields=[
                ExtractedField(
                    name="pay", value="1200.50",
                    source_document=synthetic_source_pdf.name, source_page=1,
                    certainty=Certainty.HIGH,
                ),
            ],
        )
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        pay_val = ws.cell(row=2, column=header_map["Pay"]).value

        assert pay_val == pytest.approx(1200.50)

    def test_unparseable_pay_stays_as_string(self, synthetic_source_pdf: Path, tmp_path: Path) -> None:
        """When pay cannot be parsed as float it is kept as a string rather than crashing."""
        result = make_extraction_result(
            synthetic_source_pdf,
            fields=[
                ExtractedField(
                    name="pay", value="N/A",
                    source_document=synthetic_source_pdf.name, source_page=1,
                    certainty=Certainty.REVIEW,
                ),
            ],
        )
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        pay_val = ws.cell(row=2, column=header_map["Pay"]).value

        assert pay_val == "N/A"

    def test_raw_pay_with_currency_symbol_parsed_to_float(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        """Raw LLM value '$1,500.00' must be written as float 1500.0, not as a string."""
        result = make_extraction_result(
            synthetic_source_pdf,
            fields=[
                ExtractedField(
                    name="pay", value="$1,500.00",
                    source_document=synthetic_source_pdf.name, source_page=1,
                    certainty=Certainty.HIGH,
                ),
            ],
        )
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        pay_val = ws.cell(row=2, column=header_map["Pay"]).value

        assert pay_val == pytest.approx(1500.0)

    def test_raw_pay_without_dollar_sign_but_with_commas_parsed(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        """'1,200.50' (no dollar sign, has commas) is parsed to 1200.5."""
        result = make_extraction_result(
            synthetic_source_pdf,
            fields=[
                ExtractedField(
                    name="pay", value="1,200.50",
                    source_document=synthetic_source_pdf.name, source_page=1,
                    certainty=Certainty.HIGH,
                ),
            ],
        )
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        pay_val = ws.cell(row=2, column=header_map["Pay"]).value

        assert pay_val == pytest.approx(1200.50)


class TestDuplicateRows:
    """Excluded duplicates appear as their own grayed rows naming the kept original."""

    def _load_ws(self, tmp_path: Path, results, offsets, dup_map=None):
        out = tmp_path / "out.xlsx"
        build_excel(results, out, offsets, duplicate_map=dup_map)
        return openpyxl.load_workbook(str(out)).active

    def _find_row(self, ws, doc_name: str) -> int | None:
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == doc_name:
                return row
        return None

    def test_kept_document_has_no_duplicate_note(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        """The kept original must not carry any note about its duplicates."""
        result = make_extraction_result(synthetic_source_pdf)
        dup_map = {synthetic_source_pdf.name: ["settlement (1).pdf"]}
        ws = self._load_ws(
            tmp_path, [result], {synthetic_source_pdf.name: 1}, dup_map=dup_map
        )
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        kept_row = self._find_row(ws, synthetic_source_pdf.name)
        notes_val = ws.cell(row=kept_row, column=header_map["Notes"]).value
        assert notes_val is None or notes_val == ""

    def test_single_duplicate_gets_its_own_row_naming_original(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        result = make_extraction_result(synthetic_source_pdf)
        dup_map = {synthetic_source_pdf.name: ["settlement (1).pdf"]}
        ws = self._load_ws(
            tmp_path, [result], {synthetic_source_pdf.name: 1}, dup_map=dup_map
        )
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        dup_row = self._find_row(ws, "settlement (1).pdf")

        assert dup_row is not None, "Duplicate filename did not get its own row"
        notes_val = ws.cell(row=dup_row, column=header_map["Notes"]).value
        assert notes_val == f"Exact duplicate of {synthetic_source_pdf.name}"

    def test_duplicate_row_fields_are_na_blank_and_gray(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        result = make_extraction_result(synthetic_source_pdf)
        dup_map = {synthetic_source_pdf.name: ["copy.pdf"]}
        ws = self._load_ws(
            tmp_path, [result], {synthetic_source_pdf.name: 1}, dup_map=dup_map
        )
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        dup_row = self._find_row(ws, "copy.pdf")

        # PDF Page and Certainty show NA.
        for col_name in ("PDF Page", "Certainty"):
            cell = ws.cell(row=dup_row, column=header_map[col_name])
            assert cell.value == "NA", f"{col_name} should be NA, got {cell.value!r}"
            assert cell.fill.start_color.rgb == "00D9D9D9", f"{col_name} should be gray"

        # Date and Pay are blank so the totals formulas ignore them, but stay gray.
        for col_name in ("Date", "Pay"):
            cell = ws.cell(row=dup_row, column=header_map[col_name])
            assert cell.value is None, f"{col_name} should be blank, got {cell.value!r}"
            assert cell.fill.start_color.rgb == "00D9D9D9", f"{col_name} should be gray"

    def test_multiple_duplicates_each_get_a_row(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        result = make_extraction_result(synthetic_source_pdf)
        dup_map = {
            synthetic_source_pdf.name: ["copy_a.pdf", "copy_b.pdf", "copy_c.pdf"]
        }
        ws = self._load_ws(
            tmp_path, [result], {synthetic_source_pdf.name: 1}, dup_map=dup_map
        )
        for name in ("copy_a.pdf", "copy_b.pdf", "copy_c.pdf"):
            assert self._find_row(ws, name) is not None, f"{name} has no row"

    def test_duplicate_row_contributes_nothing_to_pay_totals(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        """The dup row sits above the bottom TOTALS row with a blank Pay cell."""
        result = make_extraction_result(synthetic_source_pdf)
        dup_map = {synthetic_source_pdf.name: ["copy.pdf"]}
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 1}, duplicate_map=dup_map)
        ws = openpyxl.load_workbook(str(out), data_only=False).active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

        # Layout: row 2 payment, row 3 duplicate, row 4 TOTALS (the very bottom).
        totals_row = self._find_row(ws, "TOTALS")
        assert totals_row == ws.max_row, "TOTALS must be the very last row"
        dup_row = self._find_row(ws, "copy.pdf")
        assert dup_row < totals_row, "duplicate row must sit above the TOTALS row"

        # The dup row's Pay cell is blank, so it cannot affect the SUM.
        assert ws.cell(row=dup_row, column=header_map["Pay"]).value is None
        # SUM spans the full data range; blank dup cell is ignored by Excel.
        pay_formula = ws.cell(row=totals_row, column=header_map["Pay"]).value
        assert f"E2:E{totals_row - 1}" in pay_formula, (
            f"SUM must span the data rows, got {pay_formula!r}"
        )


class TestNonFinancialRows:
    """Non-payment documents are grayed out with NA fields and excluded from totals."""

    def _find_row(self, ws, doc_name: str) -> int | None:
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == doc_name:
                return row
        return None

    def test_non_financial_doc_is_grayed_with_na_blank_and_note(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        result = make_extraction_result(synthetic_source_pdf)
        result.is_payment_document = False
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        row = self._find_row(ws, synthetic_source_pdf.name)

        assert row is not None
        assert ws.cell(row=row, column=header_map["Notes"]).value == _NONFINANCIAL_NOTE
        for col_name in ("PDF Page", "Certainty"):
            cell = ws.cell(row=row, column=header_map[col_name])
            assert cell.value == "NA"
            assert cell.fill.start_color.rgb == "00D9D9D9"
        for col_name in ("Date", "Pay"):
            cell = ws.cell(row=row, column=header_map[col_name])
            assert cell.value is None
            assert cell.fill.start_color.rgb == "00D9D9D9"

    def test_non_financial_doc_contributes_nothing_to_pay_totals(
        self, synthetic_source_pdf: Path, synthetic_source_pdf_b: Path, tmp_path: Path
    ) -> None:
        """A non-payment doc has a blank Pay cell and sits above the bottom TOTALS row."""
        payment = make_extraction_result(synthetic_source_pdf, content_hash="h1")
        nonpayment = make_extraction_result(synthetic_source_pdf_b, content_hash="h2")
        nonpayment.is_payment_document = False
        out = tmp_path / "out.xlsx"
        build_excel([payment, nonpayment], out, {synthetic_source_pdf.name: 1})

        ws = openpyxl.load_workbook(str(out), data_only=False).active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        totals_row = self._find_row(ws, "TOTALS")
        assert totals_row == ws.max_row, "TOTALS must be the very last row"

        nonpayment_row = self._find_row(ws, synthetic_source_pdf_b.name)
        assert nonpayment_row < totals_row, "non-payment row must sit above TOTALS"
        assert ws.cell(row=nonpayment_row, column=header_map["Pay"]).value is None

        pay_formula = ws.cell(row=totals_row, column=header_map["Pay"]).value
        assert f"E2:E{totals_row - 1}" in pay_formula


class TestNoValuesReviewRow:
    """A payment doc with nothing extracted gets a red review row with a page range."""

    def _no_values_result(self, source_path: Path) -> DocumentExtractionResult:
        return DocumentExtractionResult(
            source_path=source_path,
            content_hash="novals",
            page_count=6,
            loads=[ExtractedLoad(index=1, pay=None, date=None)],
            is_payment_document=True,
        )

    def test_review_note_and_red_certainty(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        result = self._no_values_result(synthetic_source_pdf)
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 3}, {synthetic_source_pdf.name: 6})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

        assert ws.cell(row=2, column=header_map["Certainty"]).value == "Review"
        assert ws.cell(row=2, column=header_map["Certainty"]).fill.start_color.rgb == "00FFC7CE"
        assert ws.cell(row=2, column=header_map["Notes"]).value == _REVIEW_NO_VALUES_NOTE

    def test_pdf_page_shows_inclusive_range(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        """A doc starting at page 3 and spanning 6 pages must show '3 - 8'."""
        result = self._no_values_result(synthetic_source_pdf)
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 3}, {synthetic_source_pdf.name: 6})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        assert ws.cell(row=2, column=header_map["PDF Page"]).value == "3 - 8"

    def test_single_page_doc_shows_single_number(
        self, synthetic_source_pdf: Path, tmp_path: Path
    ) -> None:
        result = self._no_values_result(synthetic_source_pdf)
        result.page_count = 1
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 4}, {synthetic_source_pdf.name: 1})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        assert ws.cell(row=2, column=header_map["PDF Page"]).value == "4"


class TestTierOrdering:
    """Rows are ordered: dated payment docs, then undated, then excluded docs."""

    def test_undated_payment_doc_sorts_below_dated(
        self, synthetic_source_pdf: Path, synthetic_source_pdf_b: Path, tmp_path: Path
    ) -> None:
        from tests.unit.report.conftest import make_load

        dated = make_extraction_result(
            synthetic_source_pdf,
            content_hash="h1",
            loads=[make_load(synthetic_source_pdf, date_value="03/01/2024")],
        )
        undated = make_extraction_result(
            synthetic_source_pdf_b,
            content_hash="h2",
            loads=[make_load(synthetic_source_pdf_b, date_value=None)],
        )
        out = tmp_path / "out.xlsx"
        # Pass undated first to prove ordering is by tier, not input order.
        build_excel([undated, dated], out, {})

        ws = openpyxl.load_workbook(str(out)).active
        assert ws.cell(row=2, column=1).value == synthetic_source_pdf.name
        assert ws.cell(row=3, column=1).value == synthetic_source_pdf_b.name

    def test_excluded_docs_sort_to_the_bottom(
        self, synthetic_source_pdf: Path, synthetic_source_pdf_b: Path, tmp_path: Path
    ) -> None:
        payment = make_extraction_result(synthetic_source_pdf, content_hash="h1")
        nonpayment = make_extraction_result(synthetic_source_pdf_b, content_hash="h2")
        nonpayment.is_payment_document = False
        out = tmp_path / "out.xlsx"
        # Non-payment passed first; it must still land below the payment row.
        build_excel([nonpayment, payment], out, {synthetic_source_pdf.name: 1})

        ws = openpyxl.load_workbook(str(out)).active
        assert ws.cell(row=2, column=1).value == synthetic_source_pdf.name
        totals_row = next(
            row
            for row in range(2, ws.max_row + 1)
            if ws.cell(row=row, column=1).value == "TOTALS"
        )
        nonpayment_row = next(
            row
            for row in range(2, ws.max_row + 1)
            if ws.cell(row=row, column=1).value == synthetic_source_pdf_b.name
        )
        # Excluded rows sit below the payment row but above the bottom TOTALS row.
        assert 2 < nonpayment_row < totals_row
        assert totals_row == ws.max_row


class TestDateFormatNormalization:
    """Dates from all newly-supported surface forms are written as datetime objects.

    Each case represents a real pattern seen in OCR fixtures that the old
    parser could not handle.  They must now produce a native Excel date so
    the MAX/MIN totals formula is reliable.
    """

    @pytest.mark.parametrize(
        "raw_date, expected_dt",
        [
            ("3/19/2024",                datetime.datetime(2024, 3, 19)),   # single-digit month
            ("10/29/24",                 datetime.datetime(2024, 10, 29)),  # two-digit year
            ("2024-03-13",               datetime.datetime(2024, 3, 13)),   # ISO
            ("March 13, 2024 (Wed)",     datetime.datetime(2024, 3, 13)),   # trailing weekday
            ("Mar 11, 2024 at 11:52 AM", datetime.datetime(2024, 3, 11)),   # trailing time
        ],
        ids=[
            "single_digit_month",
            "two_digit_year",
            "iso_format",
            "trailing_weekday",
            "trailing_time",
        ],
    )
    def test_dates_normalize_to_excel_datetime(
        self,
        raw_date: str,
        expected_dt: datetime.datetime,
        synthetic_source_pdf: Path,
        tmp_path: Path,
    ) -> None:
        result = make_extraction_result(
            synthetic_source_pdf,
            fields=[
                ExtractedField(
                    name="date", value=raw_date,
                    source_document=synthetic_source_pdf.name, source_page=1,
                    certainty=Certainty.HIGH,
                ),
                ExtractedField(
                    name="pay", value="500.00",
                    source_document=synthetic_source_pdf.name, source_page=1,
                    certainty=Certainty.HIGH,
                ),
            ],
        )
        out = tmp_path / "out.xlsx"
        build_excel([result], out, {synthetic_source_pdf.name: 2})

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        cell_value = ws.cell(row=2, column=header_map["Date"]).value

        assert cell_value == expected_dt, (
            f"Date '{raw_date}' -> expected {expected_dt}, got {cell_value!r}"
        )
