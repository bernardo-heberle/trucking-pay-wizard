"""Unit tests for date-based sorting in report assembly."""

from __future__ import annotations

from pathlib import Path

import fitz
import openpyxl
import pytest

from src.extract.models import DocumentExtractionResult, ExtractedField, ExtractedLoad, SourceSpan
from src.ocr.models import BoundingBox
from src.report import build_report


def _make_result(
    source_path: Path,
    date_value: str | None,
    pay_value: str = "500.00",
    content_hash: str = "a" * 64,
) -> DocumentExtractionResult:
    pay = ExtractedField(
        name="pay", value=pay_value,
        source_document=source_path.name, source_page=1,
        source_spans=[SourceSpan(page_number=1, bounding_box=BoundingBox(x=1, y=1, width=2, height=0.25))],
    )
    date = (
        ExtractedField(
            name="date", value=date_value,
            source_document=source_path.name, source_page=1,
            source_spans=[SourceSpan(page_number=1, bounding_box=BoundingBox(x=1, y=2, width=2, height=0.25))],
        )
        if date_value is not None
        else None
    )
    return DocumentExtractionResult(
        source_path=source_path,
        content_hash=content_hash,
        loads=[ExtractedLoad(index=1, pay=pay, date=date)],
        page_count=1,
    )


def _make_source_pdf(path: Path) -> Path:
    doc = fitz.open()
    page = doc.new_page(width=612, height=792)
    page.insert_text((72, 72), f"Source: {path.name}")
    doc.save(str(path))
    doc.close()
    return path


class TestDateSortedReportOrder:

    def test_excel_rows_in_chronological_order(self, tmp_path: Path) -> None:
        """Documents with earlier dates appear in earlier rows."""
        pdf_a = _make_source_pdf(tmp_path / "doc_late.pdf")
        pdf_b = _make_source_pdf(tmp_path / "doc_early.pdf")

        late = _make_result(pdf_a, "06/30/2024", content_hash="b" * 64)
        early = _make_result(pdf_b, "01/15/2024", content_hash="c" * 64)

        out_dir = tmp_path / "output"
        _, excel_path = build_report([late, early], out_dir)

        wb = openpyxl.load_workbook(str(excel_path))
        ws = wb.active
        row2_doc = ws.cell(row=2, column=1).value
        row3_doc = ws.cell(row=3, column=1).value
        assert row2_doc == "doc_early.pdf"
        assert row3_doc == "doc_late.pdf"

    def test_pdf_source_pages_in_chronological_order(self, tmp_path: Path) -> None:
        """Source pages appear earliest-first: the early doc's page precedes the late doc's page."""
        pdf_a = _make_source_pdf(tmp_path / "doc_late.pdf")
        pdf_b = _make_source_pdf(tmp_path / "doc_early.pdf")

        late = _make_result(pdf_a, "12/25/2024", content_hash="b" * 64)
        early = _make_result(pdf_b, "03/01/2024", content_hash="c" * 64)

        out_dir = tmp_path / "output"
        pdf_path, _ = build_report([late, early], out_dir)

        doc = fitz.open(str(pdf_path))
        # Each source PDF embeds its filename as text; verify page order.
        page0_text = doc[0].get_text()
        page1_text = doc[1].get_text()
        doc.close()

        assert "doc_early.pdf" in page0_text
        assert "doc_late.pdf" in page1_text

    def test_month_name_format_sorted_correctly(self, tmp_path: Path) -> None:
        """'Month D, YYYY' format sorts correctly against 'MM/DD/YYYY'."""
        pdf_a = _make_source_pdf(tmp_path / "v2dispatch.pdf")
        pdf_b = _make_source_pdf(tmp_path / "central.pdf")

        v2 = _make_result(pdf_a, "March 20, 2024", content_hash="b" * 64)
        cd = _make_result(pdf_b, "03/09/2024", content_hash="c" * 64)

        out_dir = tmp_path / "output"
        _, excel_path = build_report([v2, cd], out_dir)

        wb = openpyxl.load_workbook(str(excel_path))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "central.pdf"
        assert ws.cell(row=3, column=1).value == "v2dispatch.pdf"

    def test_no_date_sorts_to_end(self, tmp_path: Path) -> None:
        """A document with no date field appears after all dated documents."""
        pdf_dated = _make_source_pdf(tmp_path / "dated.pdf")
        pdf_nodates = _make_source_pdf(tmp_path / "nodate.pdf")

        dated = _make_result(pdf_dated, "01/01/2024", content_hash="b" * 64)
        no_date = _make_result(pdf_nodates, None, content_hash="c" * 64)

        out_dir = tmp_path / "output"
        _, excel_path = build_report([no_date, dated], out_dir)

        wb = openpyxl.load_workbook(str(excel_path))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "dated.pdf"
        assert ws.cell(row=3, column=1).value == "nodate.pdf"

    def test_same_date_preserves_original_order(self, tmp_path: Path) -> None:
        """Documents with the same date keep their input order."""
        pdf_a = _make_source_pdf(tmp_path / "alpha.pdf")
        pdf_b = _make_source_pdf(tmp_path / "bravo.pdf")

        a = _make_result(pdf_a, "05/15/2024", content_hash="b" * 64)
        b = _make_result(pdf_b, "05/15/2024", content_hash="c" * 64)

        out_dir = tmp_path / "output"
        _, excel_path = build_report([a, b], out_dir)

        wb = openpyxl.load_workbook(str(excel_path))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "alpha.pdf"
        assert ws.cell(row=3, column=1).value == "bravo.pdf"
